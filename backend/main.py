from fastapi import FastAPI, UploadFile, File, Form, Depends, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from backend.config import settings
from backend.models.models import Base, Profe, Estudiante, Curso, Quiz, Resultado, CursoEstudiante
from backend.services.suscripcion_service import (
    profe_activo, estudiante_activo, activar_profe, activar_estudiante,
    desactivar_profe, desactivar_estudiante
)
from backend.services.pdf_service import procesar_pdf_zipgrade
import uuid, os, httpx, io
import boto3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np
from datetime import datetime, timedelta

app = FastAPI(title="ZipGrade System API", version="1.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

engine = create_engine(settings.DATABASE_URL)
Base.metadata.create_all(bind=engine)
SessionLocal = sessionmaker(bind=engine)

BOT_PROFE_TOKEN = os.getenv("BOT_PROFE_TOKEN", "")
BOT_ESTUDIANTE_TOKEN = os.getenv("BOT_ESTUDIANTE_TOKEN", "")
BASE_URL = os.getenv("RAILWAY_PUBLIC_DOMAIN", "")
R2_ACCOUNT_ID = os.getenv("R2_ACCOUNT_ID", "")
R2_ACCESS_KEY = os.getenv("R2_ACCESS_KEY", "")
R2_SECRET_KEY = os.getenv("R2_SECRET_KEY", "")
R2_BUCKET_NAME = os.getenv("R2_BUCKET_NAME", "zipgrade-pdfs")
R2_PUBLIC_URL = os.getenv("R2_PUBLIC_URL", "")
PROFE_CHAT_ID = 8911705192

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def subir_pdf_r2(pdf_bytes: bytes, nombre_archivo: str) -> str:
    try:
        client = boto3.client(
            "s3",
            endpoint_url=f"https://{R2_ACCOUNT_ID}.r2.cloudflarestorage.com",
            aws_access_key_id=R2_ACCESS_KEY,
            aws_secret_access_key=R2_SECRET_KEY,
            region_name="auto"
        )
        client.put_object(
            Bucket=R2_BUCKET_NAME,
            Key=nombre_archivo,
            Body=pdf_bytes,
            ContentType="application/pdf"
        )
        return f"{R2_PUBLIC_URL}/{nombre_archivo}"
    except Exception as e:
        print(f"Error subiendo PDF a R2: {e}")
        return ""

def generar_grafico_estudiante(nombre_est, resultados):
    if not resultados:
        return None
    materias = {}
    for r in resultados:
        curso = r.curso_nombre or "Sin curso"
        quiz = r.quiz_nombre or "Quiz"
        nota = float(r.nota) if r.nota else 0
        porcentaje = float(r.porcentaje) if r.porcentaje else 0
        key = f"{curso}\n{quiz}"
        materias[key] = {"nota": nota, "porcentaje": porcentaje}
    etiquetas = list(materias.keys())
    notas = [materias[k]["nota"] for k in etiquetas]
    porcentajes = [materias[k]["porcentaje"] for k in etiquetas]
    x = np.arange(len(etiquetas))
    ancho = 0.35
    fig, ax1 = plt.subplots(figsize=(max(10, len(etiquetas) * 1.8), 7))
    fig.patch.set_facecolor('#F8F9FA')
    ax1.set_facecolor('#F8F9FA')
    colores = ['#27AE60' if n >= 3.5 else '#F39C12' if n >= 3.0 else '#E74C3C' for n in notas]
    bars1 = ax1.bar(x - ancho/2, notas, ancho, color=colores, alpha=0.85, edgecolor='white', linewidth=1.5)
    ax2 = ax1.twinx()
    bars2 = ax2.bar(x + ancho/2, porcentajes, ancho, color='#3498DB', alpha=0.6, edgecolor='white', linewidth=1.5)
    ax1.axhline(y=3.0, color='#E74C3C', linestyle='--', linewidth=1.5, alpha=0.7)
    for bar, nota in zip(bars1, notas):
        ax1.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.05,
                f'{nota:.1f}', ha='center', va='bottom', fontsize=9, fontweight='bold')
    for bar, pct in zip(bars2, porcentajes):
        ax2.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5,
                f'{pct:.0f}%', ha='center', va='bottom', fontsize=9, color='#2980B9', fontweight='bold')
    ax1.set_xlabel('Materia / Quiz', fontsize=11, fontweight='bold')
    ax1.set_ylabel('Nota (sobre 5.0)', fontsize=11, fontweight='bold')
    ax2.set_ylabel('Porcentaje (%)', fontsize=11, fontweight='bold', color='#2980B9')
    ax1.set_title(f'Rendimiento academico\n{nombre_est}', fontsize=13, fontweight='bold', pad=15)
    ax1.set_xticks(x)
    ax1.set_xticklabels(etiquetas, fontsize=8)
    ax1.set_ylim(0, 6)
    ax2.set_ylim(0, 120)
    verde = mpatches.Patch(color='#27AE60', label='Aprobado (>=3.5)')
    amarillo = mpatches.Patch(color='#F39C12', label='Aprobado (3.0-3.4)')
    rojo = mpatches.Patch(color='#E74C3C', label='Reprobado (<3.0)')
    azul = mpatches.Patch(color='#3498DB', alpha=0.6, label='Porcentaje %')
    ax1.legend(handles=[verde, amarillo, rojo, azul], loc='upper right', fontsize=8)
    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    buf.seek(0)
    plt.close()
    return buf.getvalue()

def generar_grafico_profe(resultados_todos):
    if not resultados_todos:
        return None
    grupos = {}
    for r in resultados_todos:
        curso = r.curso_nombre or "Sin curso"
        quiz = r.quiz_nombre or "Quiz"
        nota = float(r.nota) if r.nota else 0
        key = f"{curso}\n{quiz}"
        if key not in grupos:
            grupos[key] = {"aprobados": 0, "reprobados": 0, "total": 0}
        grupos[key]["total"] += 1
        if nota >= 3.0:
            grupos[key]["aprobados"] += 1
        else:
            grupos[key]["reprobados"] += 1
    etiquetas = list(grupos.keys())
    aprobados = [grupos[k]["aprobados"] for k in etiquetas]
    reprobados = [grupos[k]["reprobados"] for k in etiquetas]
    pct_reprobados = [round(grupos[k]["reprobados"] / grupos[k]["total"] * 100, 1) for k in etiquetas]
    x = np.arange(len(etiquetas))
    ancho = 0.35
    fig, ax = plt.subplots(figsize=(max(10, len(etiquetas) * 2), 7))
    fig.patch.set_facecolor('#F8F9FA')
    ax.set_facecolor('#F8F9FA')
    bars1 = ax.bar(x - ancho/2, aprobados, ancho, label='Aprobados', color='#27AE60', alpha=0.85, edgecolor='white', linewidth=1.5)
    bars2 = ax.bar(x + ancho/2, reprobados, ancho, label='Reprobados', color='#E74C3C', alpha=0.85, edgecolor='white', linewidth=1.5)
    for bar, val in zip(bars1, aprobados):
        if val > 0:
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.1,
                    str(val), ha='center', va='bottom', fontsize=10, fontweight='bold', color='#27AE60')
    for bar, val, pct in zip(bars2, reprobados, pct_reprobados):
        if val > 0:
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.1,
                    f'{val}\n({pct}%)', ha='center', va='bottom', fontsize=9, fontweight='bold', color='#C0392B')
    ax.set_xlabel('Materia / Quiz', fontsize=11, fontweight='bold')
    ax.set_ylabel('Numero de estudiantes', fontsize=11, fontweight='bold')
    ax.set_title('Estadisticas del grupo\nAprobados vs Reprobados', fontsize=13, fontweight='bold', pad=15)
    ax.set_xticks(x)
    ax.set_xticklabels(etiquetas, fontsize=8)
    ax.legend(fontsize=10)
    ax.yaxis.set_major_locator(plt.MaxNLocator(integer=True))
    plt.tight_layout()
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    buf.seek(0)
    plt.close()
    return buf.getvalue()

def generar_excel(resultados, titulo):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notas"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    title_font = Font(bold=True, size=13, color="1F4E79")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A1:D1")
    ws["A1"] = titulo
    ws["A1"].font = title_font
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 25
    encabezados = ["#", "Estudiante", "Nota (sobre 5.0)", "Porcentaje"]
    anchos = [5, 30, 18, 15]
    for col, (h, ancho) in enumerate(zip(encabezados, anchos), 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[cell.column_letter].width = ancho
    ws.row_dimensions[2].height = 20
    for i, r in enumerate(resultados, 1):
        nota = float(r.nota) if r.nota else 0
        porcentaje = float(r.porcentaje) if r.porcentaje else 0
        fila = i + 2
        ws.cell(row=fila, column=1, value=i).alignment = center
        ws.cell(row=fila, column=2, value=r.nombre_temp or "").alignment = left
        ws.cell(row=fila, column=3, value=f"{nota:.2f} / 5.0").alignment = center
        ws.cell(row=fila, column=4, value=f"{porcentaje:.1f}%").alignment = center
        nota_cell = ws.cell(row=fila, column=3)
        if nota >= 3.5:
            nota_cell.fill = PatternFill("solid", fgColor="C6EFCE")
            nota_cell.font = Font(color="276221", bold=True)
        elif nota >= 3.0:
            nota_cell.fill = PatternFill("solid", fgColor="FFEB9C")
            nota_cell.font = Font(color="9C5700", bold=True)
        else:
            nota_cell.fill = PatternFill("solid", fgColor="FFC7CE")
            nota_cell.font = Font(color="9C0006", bold=True)
        ws.row_dimensions[fila].height = 18
    total = len(resultados)
    if total > 0:
        promedio = sum(float(r.nota) for r in resultados if r.nota) / total
        aprobados = sum(1 for r in resultados if r.nota and float(r.nota) >= 3.0)
        fila_prom = total + 4
        ws.merge_cells(f"A{fila_prom}:D{fila_prom}")
        ws[f"A{fila_prom}"] = f"Total: {total}  |  Aprobados: {aprobados}  |  Reprobados: {total - aprobados}"
        ws[f"A{fila_prom}"].font = Font(bold=True, color="1F4E79")
        ws[f"A{fila_prom+1}"] = "Promedio del grupo:"
        ws[f"A{fila_prom+1}"].font = Font(bold=True)
        ws[f"C{fila_prom+1}"] = f"{promedio:.2f} / 5.0"
        ws[f"C{fila_prom+1}"].font = Font(bold=True, color="1F4E79")
        ws[f"C{fila_prom+1}"].alignment = Alignment(horizontal="center")
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

async def send_message(token, chat_id, text, reply_markup=None):
    payload = {"chat_id": chat_id, "text": text, "parse_mode": "HTML"}
    if reply_markup:
        payload["reply_markup"] = reply_markup
    async with httpx.AsyncClient() as client:
        await client.post(f"https://api.telegram.org/bot{token}/sendMessage", json=payload)

async def send_photo(token, chat_id, photo_url, caption=""):
    async with httpx.AsyncClient() as client:
        await client.post(f"https://api.telegram.org/bot{token}/sendPhoto",
            json={"chat_id": chat_id, "photo": photo_url, "caption": caption})

async def send_photo_bytes(token, chat_id, photo_bytes, caption=""):
    async with httpx.AsyncClient(timeout=60) as client:
        await client.post(f"https://api.telegram.org/bot{token}/sendPhoto",
            data={"chat_id": chat_id, "caption": caption},
            files={"photo": ("grafico.png", photo_bytes, "image/png")})

async def send_document_url(token, chat_id, doc_url, caption=""):
    async with httpx.AsyncClient() as client:
        await client.post(f"https://api.telegram.org/bot{token}/sendDocument",
            json={"chat_id": chat_id, "document": doc_url, "caption": caption})

async def send_excel(token, chat_id, excel_bytes, filename, caption=""):
    async with httpx.AsyncClient(timeout=60) as client:
        await client.post(f"https://api.telegram.org/bot{token}/sendDocument",
            data={"chat_id": chat_id, "caption": caption},
            files={"document": (filename, excel_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})

async def send_voice(token, chat_id, file_id, source_token, caption=""):
    try:
        async with httpx.AsyncClient(timeout=60) as client:
            r = await client.get(f"https://api.telegram.org/bot{source_token}/getFile",
                params={"file_id": file_id})
            file_path = r.json()["result"]["file_path"]
            audio_r = await client.get(f"https://api.telegram.org/file/bot{source_token}/{file_path}")
            audio_bytes = audio_r.content
            await client.post(f"https://api.telegram.org/bot{token}/sendVoice",
                data={"chat_id": chat_id, "caption": caption},
                files={"voice": ("voice.ogg", audio_bytes, "audio/ogg")})
    except Exception as e:
        print(f"Error enviando nota de voz: {e}")

async def reenviar_archivo(token_destino, chat_id_destino, file_id, token_origen, file_name="archivo", caption="", tipo="document"):
    try:
        async with httpx.AsyncClient(timeout=120) as client:
            r = await client.get(f"https://api.telegram.org/bot{token_origen}/getFile",
                params={"file_id": file_id})
            file_path = r.json()["result"]["file_path"]
            file_r = await client.get(f"https://api.telegram.org/file/bot{token_origen}/{file_path}")
            file_bytes = file_r.content
            if tipo == "photo":
                await client.post(f"https://api.telegram.org/bot{token_destino}/sendPhoto",
                    data={"chat_id": chat_id_destino, "caption": caption},
                    files={"photo": (file_name, file_bytes, "image/jpeg")})
            elif tipo == "video":
                await client.post(f"https://api.telegram.org/bot{token_destino}/sendVideo",
                    data={"chat_id": chat_id_destino, "caption": caption},
                    files={"video": (file_name, file_bytes, "video/mp4")})
            else:
                await client.post(f"https://api.telegram.org/bot{token_destino}/sendDocument",
                    data={"chat_id": chat_id_destino, "caption": caption},
                    files={"document": (file_name, file_bytes, "application/octet-stream")})
    except Exception as e:
        print(f"Error reenviando archivo: {e}")

# ─── HELPERS DE GRUPO VIRTUAL ────────────────────────────────────────────────

def get_grupo_activo(db, telegram_id):
    r = db.query(Resultado).filter(
        Resultado.nombre_temp == f"__grupo__{telegram_id}__curso_id"
    ).first()
    return r.quiz_nombre if r else None

def get_grupo_nombre(db, telegram_id):
    r = db.query(Resultado).filter(
        Resultado.nombre_temp == f"__grupo__{telegram_id}__curso_nombre"
    ).first()
    return r.quiz_nombre if r else ""

def entrar_grupo(db, telegram_id, curso_id, curso_nombre):
    for clave, valor in [("curso_id", str(curso_id)), ("curso_nombre", curso_nombre)]:
        r = db.query(Resultado).filter(
            Resultado.nombre_temp == f"__grupo__{telegram_id}__{clave}"
        ).first()
        if r:
            r.quiz_nombre = valor
        else:
            db.add(Resultado(id=uuid.uuid4(),
                nombre_temp=f"__grupo__{telegram_id}__{clave}",
                quiz_nombre=valor, confirmado=False))
    db.commit()

def salir_grupo(db, telegram_id):
    for clave in ["curso_id", "curso_nombre"]:
        db.query(Resultado).filter(
            Resultado.nombre_temp == f"__grupo__{telegram_id}__{clave}"
        ).delete(synchronize_session=False)
    db.commit()

async def transmitir_grupo(db, curso_id, remitente_id, remitente_nombre, es_profe,
                            tipo, file_id=None, text=None, file_name="archivo"):
    """
    Envía el mensaje/archivo a todos los miembros del grupo.
    - Si envía un ESTUDIANTE: llega a todos los compañeros inscritos + al profe del curso (siempre).
    - Si envía el PROFE: llega a todos los estudiantes inscritos.
    """
    prefijo = f"👨‍🏫 <b>[PROFE] {remitente_nombre}:</b>" if es_profe else f"👥 <b>[GRUPO] {remitente_nombre}:</b>"

    curso = db.query(Curso).filter(Curso.id == curso_id).first()
    inscripciones = db.query(CursoEstudiante).filter(
        CursoEstudiante.curso_id == curso_id
    ).all()

    destinatarios = []  # (token, chat_id)

    # Todos los estudiantes inscritos excepto el remitente
    for ins in inscripciones:
        est = db.query(Estudiante).filter(Estudiante.id == ins.estudiante_id).first()
        if est and est.telegram_id != remitente_id:
            destinatarios.append((BOT_ESTUDIANTE_TOKEN, est.telegram_id))

    # El profe del curso SIEMPRE recibe cuando un estudiante envía algo
    # (sin importar si está o no inmerso en el grupo)
    if not es_profe and curso:
        profe_curso = db.query(Profe).filter(Profe.id == curso.profe_id).first()
        if profe_curso and profe_curso.telegram_id != remitente_id:
            destinatarios.append((BOT_PROFE_TOKEN, profe_curso.telegram_id))

    origen_token = BOT_PROFE_TOKEN if es_profe else BOT_ESTUDIANTE_TOKEN

    for token, chat_id in destinatarios:
        try:
            async with httpx.AsyncClient(timeout=30) as client:
                if tipo == "text":
                    await client.post(f"https://api.telegram.org/bot{token}/sendMessage",
                        json={"chat_id": chat_id, "text": f"{prefijo}\n{text}", "parse_mode": "HTML"})

                elif tipo == "voice":
                    await client.post(f"https://api.telegram.org/bot{token}/sendMessage",
                        json={"chat_id": chat_id, "text": f"{prefijo}\n🎙️ Nota de voz", "parse_mode": "HTML"})
                    r = await client.get(f"https://api.telegram.org/bot{origen_token}/getFile",
                        params={"file_id": file_id})
                    fp = r.json()["result"]["file_path"]
                    audio_r = await client.get(f"https://api.telegram.org/file/bot{origen_token}/{fp}")
                    await client.post(f"https://api.telegram.org/bot{token}/sendVoice",
                        data={"chat_id": chat_id},
                        files={"voice": ("voice.ogg", audio_r.content, "audio/ogg")})

                elif tipo in ("photo", "video", "document"):
                    r = await client.get(f"https://api.telegram.org/bot{origen_token}/getFile",
                        params={"file_id": file_id})
                    fp = r.json()["result"]["file_path"]
                    file_r = await client.get(f"https://api.telegram.org/file/bot{origen_token}/{fp}")
                    fb = file_r.content
                    caption = prefijo
                    if tipo == "photo":
                        await client.post(f"https://api.telegram.org/bot{token}/sendPhoto",
                            data={"chat_id": chat_id, "caption": caption, "parse_mode": "HTML"},
                            files={"photo": (file_name, fb, "image/jpeg")})
                    elif tipo == "video":
                        await client.post(f"https://api.telegram.org/bot{token}/sendVideo",
                            data={"chat_id": chat_id, "caption": caption, "parse_mode": "HTML"},
                            files={"video": (file_name, fb, "video/mp4")})
                    else:
                        await client.post(f"https://api.telegram.org/bot{token}/sendDocument",
                            data={"chat_id": chat_id, "caption": caption, "parse_mode": "HTML"},
                            files={"document": (file_name, fb, "application/octet-stream")})
        except Exception as e:
            print(f"Error transmitiendo a {chat_id}: {e}")

# ─────────────────────────────────────────────────────────────────────────────

def sesion_activa(db, telegram_id):
    r = db.query(Resultado).filter(
        Resultado.nombre_temp == f"__estado__{telegram_id}__sesion_inicio"
    ).first()
    if not r or not r.quiz_nombre:
        return False
    try:
        inicio = datetime.fromisoformat(r.quiz_nombre)
        return datetime.now() < inicio + timedelta(minutes=15)
    except:
        return False

def tiempo_restante(db, telegram_id):
    r = db.query(Resultado).filter(
        Resultado.nombre_temp == f"__estado__{telegram_id}__sesion_inicio"
    ).first()
    if not r or not r.quiz_nombre:
        return 0
    try:
        inicio = datetime.fromisoformat(r.quiz_nombre)
        restante = (inicio + timedelta(minutes=15)) - datetime.now()
        return max(0, int(restante.total_seconds() / 60))
    except:
        return 0

def get_estado(db, telegram_id, clave):
    r = db.query(Resultado).filter(
        Resultado.nombre_temp == f"__estado__{telegram_id}__{clave}"
    ).first()
    return r.quiz_nombre if r else None

def set_estado(db, telegram_id, clave, valor):
    r = db.query(Resultado).filter(
        Resultado.nombre_temp == f"__estado__{telegram_id}__{clave}"
    ).first()
    if r:
        r.quiz_nombre = valor
    else:
        db.add(Resultado(id=uuid.uuid4(),
            nombre_temp=f"__estado__{telegram_id}__{clave}",
            quiz_nombre=valor, confirmado=False))
    db.commit()

def del_estado(db, telegram_id, clave):
    db.query(Resultado).filter(
        Resultado.nombre_temp == f"__estado__{telegram_id}__{clave}"
    ).delete(synchronize_session=False)
    db.commit()

@app.on_event("startup")
async def set_webhooks():
    if BOT_PROFE_TOKEN and BASE_URL:
        async with httpx.AsyncClient() as client:
            await client.get(f"https://api.telegram.org/bot{BOT_PROFE_TOKEN}/setWebhook",
                params={"url": f"https://{BASE_URL}/webhook/profe"})
            await client.get(f"https://api.telegram.org/bot{BOT_ESTUDIANTE_TOKEN}/setWebhook",
                params={"url": f"https://{BASE_URL}/webhook/estudiante"})

@app.post("/webhook/profe")
async def webhook_profe(request: Request, db: Session = Depends(get_db)):
    data = await request.json()
    callback = data.get("callback_query", {})
    message = data.get("message", {})

    if callback:
        chat_id = callback.get("from", {}).get("id")
        telegram_id = chat_id
        cb_data = callback.get("data", "")

        if cb_data.startswith("grupo_profe_"):
            curso_id = cb_data.replace("grupo_profe_", "")
            curso = db.query(Curso).filter(Curso.id == curso_id).first()
            if curso:
                entrar_grupo(db, telegram_id, curso_id, f"{curso.nombre} {curso.grado}")
                profe_nombre = callback.get("from", {}).get("first_name", "El profe")
                await send_message(BOT_PROFE_TOKEN, chat_id,
                    f"🏫 <b>Estás en el grupo: {curso.nombre} {curso.grado}</b>\n\n"
                    f"Todo lo que escribas o envíes (texto, voz, imágenes, PDF, video) "
                    f"llegará a todos los estudiantes inscritos.\n\n"
                    f"Usa /salir_grupo para volver a tu chat normal.")
                # Notificar a estudiantes que el profe entró
                inscripciones = db.query(CursoEstudiante).filter(
                    CursoEstudiante.curso_id == curso_id
                ).all()
                for ins in inscripciones:
                    est = db.query(Estudiante).filter(Estudiante.id == ins.estudiante_id).first()
                    if est:
                        try:
                            await send_message(BOT_ESTUDIANTE_TOKEN, est.telegram_id,
                                f"🔔 <b>El profe {profe_nombre} se unió al grupo de {curso.nombre} {curso.grado}</b>\n\n"
                                f"Usa /grupos para entrar y ver lo que comparte.")
                        except:
                            pass

        elif cb_data.startswith("curso_"):
            curso_id = cb_data.replace("curso_", "")
            curso = db.query(Curso).filter(Curso.id == curso_id).first()
            if curso:
                set_estado(db, telegram_id, "curso_seleccionado", f"{curso_id}|{curso.nombre}")
                set_estado(db, telegram_id, "paso", "esperando_nombre_quiz")
                await send_message(BOT_PROFE_TOKEN, chat_id,
                    f"📚 Curso: <b>{curso.nombre} - {curso.grado}</b>\n\n✏️ Escribe el nombre del quiz:\nEjemplo: <b>Quiz 1 Primer Periodo</b>")

        elif cb_data.startswith("enviar_curso_"):
            curso_id = cb_data.replace("enviar_curso_", "")
            curso = db.query(Curso).filter(Curso.id == curso_id).first()
            if curso:
                set_estado(db, telegram_id, "enviar_archivo_curso", f"{curso_id}|{curso.nombre}")
                set_estado(db, telegram_id, "paso", "esperando_archivo_para_curso")
                await send_message(BOT_PROFE_TOKEN, chat_id,
                    f"📤 Curso: <b>{curso.nombre} - {curso.grado}</b>\n\n"
                    f"Ahora envía el archivo, imagen o video que quieres compartir con todos los estudiantes.")

        elif cb_data.startswith("excel_quiz_"):
            partes = cb_data.replace("excel_quiz_", "").split("|", 1)
            curso_buscar = partes[0]
            quiz_buscar = partes[1] if len(partes) > 1 else ""
            resultados = db.query(Resultado).filter(
                Resultado.curso_nombre.ilike(f"%{curso_buscar}%"),
                Resultado.quiz_nombre.ilike(f"%{quiz_buscar}%"),
                Resultado.confirmado == True
            ).all()
            if not resultados:
                await send_message(BOT_PROFE_TOKEN, chat_id, f"❌ No hay resultados para {quiz_buscar}.")
            else:
                titulo = f"Notas - {curso_buscar} - {quiz_buscar}"
                excel_bytes = generar_excel(resultados, titulo)
                filename = f"notas_{curso_buscar}_{quiz_buscar}.xlsx".replace(" ", "_")
                await send_excel(BOT_PROFE_TOKEN, chat_id, excel_bytes, filename,
                    f"📊 {titulo} — {len(resultados)} estudiantes")

        elif cb_data.startswith("excel_todos_"):
            curso_buscar = cb_data.replace("excel_todos_", "")
            resultados = db.query(Resultado).filter(
                Resultado.curso_nombre.ilike(f"%{curso_buscar}%"),
                Resultado.confirmado == True
            ).all()
            if not resultados:
                await send_message(BOT_PROFE_TOKEN, chat_id, f"❌ No hay resultados para {curso_buscar}.")
            else:
                titulo = f"Todas las notas - {curso_buscar}"
                excel_bytes = generar_excel(resultados, titulo)
                filename = f"notas_{curso_buscar}_todos.xlsx".replace(" ", "_")
                await send_excel(BOT_PROFE_TOKEN, chat_id, excel_bytes, filename,
                    f"📊 {titulo} — {len(resultados)} registros")

        return {"ok": True}

    chat_id = message.get("chat", {}).get("id")
    text = message.get("text", "")
    telegram_id = message.get("from", {}).get("id")
    nombre = message.get("from", {}).get("first_name", "Profe")
    document = message.get("document", {})
    voice = message.get("voice", {})
    photo = message.get("photo", [])
    video = message.get("video", {})

    if not chat_id:
        return {"ok": True}

    profe = db.query(Profe).filter(Profe.telegram_id == telegram_id).first()

    # ── MODO GRUPO ACTIVO DEL PROFE ──────────────────────────────────────────
    grupo_activo = get_grupo_activo(db, telegram_id)
    if grupo_activo and text != "/salir_grupo" and text != "/grupos":
        grupo_nombre = get_grupo_nombre(db, telegram_id)

        if voice:
            voice_file_id = voice.get("file_id")
            paso = get_estado(db, telegram_id, "paso")
            if paso and paso.startswith("responder_voz_"):
                try:
                    estudiante_dest = int(paso.replace("responder_voz_", ""))
                    del_estado(db, telegram_id, "paso")
                    await send_voice(BOT_ESTUDIANTE_TOKEN, estudiante_dest, voice_file_id, BOT_PROFE_TOKEN,
                        "🎙️ Nota de voz de tu profe")
                    await send_message(BOT_PROFE_TOKEN, chat_id, "✅ Nota de voz enviada al estudiante.")
                except Exception as e:
                    await send_message(BOT_PROFE_TOKEN, chat_id, f"❌ Error: {str(e)}")
            else:
                await transmitir_grupo(db, grupo_activo, telegram_id, nombre, True,
                    "voice", file_id=voice_file_id)
                await send_message(BOT_PROFE_TOKEN, chat_id,
                    f"🎙️ Nota de voz enviada al grupo <b>{grupo_nombre}</b>.")
            return {"ok": True}

        if photo:
            fid = photo[-1].get("file_id")
            await transmitir_grupo(db, grupo_activo, telegram_id, nombre, True,
                "photo", file_id=fid, file_name="imagen.jpg")
            await send_message(BOT_PROFE_TOKEN, chat_id,
                f"🖼️ Imagen enviada al grupo <b>{grupo_nombre}</b>.")
            return {"ok": True}

        if video:
            fid = video.get("file_id")
            fname = video.get("file_name", "video.mp4")
            await transmitir_grupo(db, grupo_activo, telegram_id, nombre, True,
                "video", file_id=fid, file_name=fname)
            await send_message(BOT_PROFE_TOKEN, chat_id,
                f"🎥 Video enviado al grupo <b>{grupo_nombre}</b>.")
            return {"ok": True}

        if document:
            fid = document.get("file_id")
            fname = document.get("file_name", "archivo")
            await transmitir_grupo(db, grupo_activo, telegram_id, nombre, True,
                "document", file_id=fid, file_name=fname)
            await send_message(BOT_PROFE_TOKEN, chat_id,
                f"📎 Archivo enviado al grupo <b>{grupo_nombre}</b>.")
            return {"ok": True}

        if text and not text.startswith("/"):
            await transmitir_grupo(db, grupo_activo, telegram_id, nombre, True,
                "text", text=text)
            return {"ok": True}
    # ─────────────────────────────────────────────────────────────────────────

    if voice:
        voice_file_id = voice.get("file_id")
        paso = get_estado(db, telegram_id, "paso")
        if paso and paso.startswith("responder_voz_"):
            try:
                estudiante_dest = int(paso.replace("responder_voz_", ""))
                del_estado(db, telegram_id, "paso")
                await send_voice(BOT_ESTUDIANTE_TOKEN, estudiante_dest, voice_file_id, BOT_PROFE_TOKEN,
                    "🎙️ Nota de voz de tu profe")
                await send_message(BOT_PROFE_TOKEN, chat_id, "✅ Nota de voz enviada al estudiante.")
            except Exception as e:
                await send_message(BOT_PROFE_TOKEN, chat_id, f"❌ Error: {str(e)}")
        else:
            await send_message(BOT_PROFE_TOKEN, chat_id,
                "❌ Primero usa <code>/responder_voz ID_ESTUDIANTE</code> y luego envía la nota de voz.")
        return {"ok": True}

    archivo_recibido = None
    archivo_tipo = None
    archivo_nombre = "archivo"
    if photo:
        archivo_recibido = photo[-1].get("file_id")
        archivo_tipo = "photo"
        archivo_nombre = "imagen.jpg"
    elif video:
        archivo_recibido = video.get("file_id")
        archivo_tipo = "video"
        archivo_nombre = video.get("file_name", "video.mp4")
    elif document and not message.get("text", ""):
        archivo_recibido = document.get("file_id")
        archivo_tipo = "document"
        archivo_nombre = document.get("file_name", "archivo")

    if archivo_recibido:
        paso = get_estado(db, telegram_id, "paso")
        if paso == "esperando_archivo_para_curso":
            curso_info = get_estado(db, telegram_id, "enviar_archivo_curso")
            if curso_info:
                curso_id, curso_nombre = curso_info.split("|", 1)
                del_estado(db, telegram_id, "paso")
                del_estado(db, telegram_id, "enviar_archivo_curso")
                inscripciones = db.query(CursoEstudiante).filter(CursoEstudiante.curso_id == curso_id).all()
                if not inscripciones:
                    await send_message(BOT_PROFE_TOKEN, chat_id,
                        f"⚠️ No hay estudiantes inscritos en <b>{curso_nombre}</b>.")
                else:
                    await send_message(BOT_PROFE_TOKEN, chat_id, f"⏳ Enviando a {len(inscripciones)} estudiantes...")
                    enviados = 0
                    for ins in inscripciones:
                        est = db.query(Estudiante).filter(Estudiante.id == ins.estudiante_id).first()
                        if est:
                            try:
                                await reenviar_archivo(BOT_ESTUDIANTE_TOKEN, est.telegram_id,
                                    archivo_recibido, BOT_PROFE_TOKEN, archivo_nombre,
                                    f"📎 Archivo de tu profe — {curso_nombre}", archivo_tipo)
                                enviados += 1
                            except:
                                pass
                    await send_message(BOT_PROFE_TOKEN, chat_id,
                        f"✅ Archivo enviado a <b>{enviados} estudiantes</b> del curso <b>{curso_nombre}</b>.")
            return {"ok": True}

        elif paso and paso.startswith("enviar_archivo_estudiante_"):
            estudiante_dest = int(paso.replace("enviar_archivo_estudiante_", ""))
            del_estado(db, telegram_id, "paso")
            try:
                await reenviar_archivo(BOT_ESTUDIANTE_TOKEN, estudiante_dest,
                    archivo_recibido, BOT_PROFE_TOKEN, archivo_nombre,
                    "📎 Archivo de tu profe", archivo_tipo)
                await send_message(BOT_PROFE_TOKEN, chat_id, "✅ Archivo enviado al estudiante.")
            except Exception as e:
                await send_message(BOT_PROFE_TOKEN, chat_id, f"❌ Error enviando archivo: {str(e)}")
            return {"ok": True}

        elif archivo_tipo == "document" and archivo_nombre.endswith(".pdf"):
            paso_actual = get_estado(db, telegram_id, "paso")
            curso_info = get_estado(db, telegram_id, "curso_seleccionado")
            quiz_nombre_estado = get_estado(db, telegram_id, "quiz_nombre")
            resultados_pendientes = db.query(Resultado).filter(
                Resultado.nombre_temp.like("PAG%"),
                Resultado.confirmado == False,
                Resultado.profe_telegram_id == telegram_id
            ).all()

            async with httpx.AsyncClient(timeout=60) as client:
                r = await client.get(f"https://api.telegram.org/bot{BOT_PROFE_TOKEN}/getFile",
                    params={"file_id": archivo_recibido})
                file_path = r.json()["result"]["file_path"]
                file_r = await client.get(f"https://api.telegram.org/file/bot{BOT_PROFE_TOKEN}/{file_path}")
                file_bytes = file_r.content

            if resultados_pendientes and paso_actual == "esperando_pdf_quiz":
                await send_message(BOT_PROFE_TOKEN, chat_id, "📄 PDF del quiz recibido. Subiendo...")
                nombre_archivo = f"quizzes/{uuid.uuid4()}.pdf"
                quiz_pdf_url = subir_pdf_r2(file_bytes, nombre_archivo)
                if quiz_pdf_url:
                    for r in resultados_pendientes:
                        r.quiz_pdf_url = quiz_pdf_url
                    db.commit()
                    set_estado(db, telegram_id, "paso", "esperando_nombres")
                    await send_message(BOT_PROFE_TOKEN, chat_id,
                        "✅ PDF del quiz guardado.\n\nAhora pega la lista de nombres:\nPAG1: Nombre Apellido\nPAG2: Nombre Apellido...")
                else:
                    await send_message(BOT_PROFE_TOKEN, chat_id, "❌ Error subiendo el PDF.")
            else:
                if not curso_info:
                    await send_message(BOT_PROFE_TOKEN, chat_id, "❌ Primero selecciona un curso con /subirquiz")
                    return {"ok": True}
                curso_id, curso_nombre = curso_info.split("|", 1)
                qnombre = quiz_nombre_estado or "Quiz"
                await send_message(BOT_PROFE_TOKEN, chat_id,
                    f"📎 PDF de ZipGrade recibido.\n📚 Curso: <b>{curso_nombre}</b>\n📝 Quiz: <b>{qnombre}</b>\n\n⏳ Procesando...")
                try:
                    resultados_lista = await procesar_pdf_zipgrade(file_bytes)
                    total = len(resultados_lista)
                    db.query(Resultado).filter(
                        Resultado.nombre_temp.like("PAG%"),
                        Resultado.confirmado == False,
                        Resultado.profe_telegram_id == telegram_id
                    ).delete(synchronize_session=False)
                    db.commit()
                    for r in resultados_lista:
                        nuevo_r = Resultado(
                            id=uuid.uuid4(),
                            nombre_temp=r["nombre"],
                            nota=r["nota"],
                            puntos=r["puntos"],
                            posibles=r["posibles"],
                            porcentaje=r["porcentaje"],
                            pagina=r.get("pagina", 0),
                            imagen_url=r.get("imagen_url", ""),
                            curso_nombre=curso_nombre,
                            quiz_nombre=qnombre,
                            profe_telegram_id=telegram_id,
                            confirmado=False
                        )
                        db.add(nuevo_r)
                    db.commit()
                    set_estado(db, telegram_id, "paso", "esperando_pdf_quiz")
                    resumen = "\n".join([f"• <b>{r['nombre']}</b>: {r['nota']}/5.0 ({r['porcentaje']}%)" for r in resultados_lista])
                    await send_message(BOT_PROFE_TOKEN, chat_id,
                        f"✅ PDF procesado: <b>{total} estudiantes</b>\n\n{resumen}\n\n"
                        f"📄 Ahora envíame el PDF del quiz (las preguntas).")
                except Exception as e:
                    await send_message(BOT_PROFE_TOKEN, chat_id, f"❌ Error procesando PDF: {str(e)}")
            return {"ok": True}

        else:
            cursos = db.query(Curso).filter(Curso.profe_id == profe.id).all() if profe else []
            if cursos:
                set_estado(db, telegram_id, "archivo_pendiente_id", archivo_recibido)
                set_estado(db, telegram_id, "archivo_pendiente_tipo", archivo_tipo)
                set_estado(db, telegram_id, "archivo_pendiente_nombre", archivo_nombre)
                set_estado(db, telegram_id, "paso", "esperando_archivo_para_curso")
                botones = {"inline_keyboard": [[{
                    "text": f"📚 {c.nombre} - {c.grado}",
                    "callback_data": f"enviar_curso_{c.id}"
                }] for c in cursos]}
                await send_message(BOT_PROFE_TOKEN, chat_id,
                    "📤 ¿A qué curso quieres enviar este archivo?",
                    reply_markup=botones)
            else:
                await send_message(BOT_PROFE_TOKEN, chat_id,
                    "❌ No tienes cursos creados. Usa /nuevocurso primero.")
            return {"ok": True}

    if text == "/start":
        if not profe:
            nuevo = Profe(id=uuid.uuid4(), telegram_id=telegram_id, nombre=nombre, email="", activo=False)
            db.add(nuevo)
            db.commit()
            await send_message(BOT_PROFE_TOKEN, chat_id,
                f"👋 Hola <b>{nombre}</b>!\n\nTu cuenta fue creada. Contacta al administrador para activar tu suscripcion.")
        else:
            if profe.activo:
                await send_message(BOT_PROFE_TOKEN, chat_id,
                    f"✅ Hola <b>{profe.nombre}</b>!\n\n📋 Comandos:\n"
                    f"/micursos - Ver tus cursos\n"
                    f"/nuevocurso - Crear un curso\n"
                    f"/subirquiz - Subir quiz\n"
                    f"/excel - Generar Excel de notas\n"
                    f"/estadisticas - Ver grafico del grupo\n"
                    f"/enviar - Enviar archivo a un curso\n"
                    f"/grupos - Entrar a un grupo virtual\n"
                    f"/estado - Ver suscripcion\n\n"
                    f"💬 Para responder individualmente:\n"
                    f"• Texto: <code>/responder ID mensaje</code>\n"
                    f"• Voz: <code>/responder_voz ID</code> y luego graba\n"
                    f"• Archivo: <code>/enviar_a ID</code> y luego envía")
            else:
                await send_message(BOT_PROFE_TOKEN, chat_id, "❌ Tu suscripcion no esta activa.")

    elif text == "/grupos":
        if not profe or not profe.activo:
            await send_message(BOT_PROFE_TOKEN, chat_id, "❌ Necesitas suscripcion activa.")
            return {"ok": True}
        cursos = db.query(Curso).filter(Curso.profe_id == profe.id).all()
        if not cursos:
            await send_message(BOT_PROFE_TOKEN, chat_id, "No tienes cursos. Usa /nuevocurso para crear uno.")
        else:
            grupo_actual = get_grupo_activo(db, telegram_id)
            grupo_nombre_actual = get_grupo_nombre(db, telegram_id)
            msg = "🏫 <b>Grupos virtuales</b>\n\nSelecciona el grupo al que quieres entrar:"
            if grupo_actual:
                msg = f"🏫 Actualmente estás en: <b>{grupo_nombre_actual}</b>\n\nUsa /salir_grupo para salir o cambia de grupo:"
            botones = {"inline_keyboard": [[{
                "text": f"🏫 {c.nombre} - {c.grado}",
                "callback_data": f"grupo_profe_{c.id}"
            }] for c in cursos]}
            await send_message(BOT_PROFE_TOKEN, chat_id, msg, reply_markup=botones)

    elif text == "/salir_grupo":
        grupo_nombre_actual = get_grupo_nombre(db, telegram_id)
        salir_grupo(db, telegram_id)
        await send_message(BOT_PROFE_TOKEN, chat_id,
            f"✅ Saliste del grupo <b>{grupo_nombre_actual}</b>.\n\nVolviste a tu chat normal de profesor.")

    elif text == "/estado":
        if profe:
            estado = "✅ Activa" if profe.activo else "❌ Inactiva"
            await send_message(BOT_PROFE_TOKEN, chat_id, f"📊 Tu suscripcion: {estado}")

    elif text == "/micursos":
        if not profe or not profe.activo:
            await send_message(BOT_PROFE_TOKEN, chat_id, "❌ Necesitas suscripcion activa.")
            return {"ok": True}
        cursos = db.query(Curso).filter(Curso.profe_id == profe.id).all()
        if not cursos:
            await send_message(BOT_PROFE_TOKEN, chat_id, "No tienes cursos. Usa /nuevocurso para crear uno.")
        else:
            lista = "\n".join([f"📚 <b>{c.nombre}</b> - {c.grado}" for c in cursos])
            await send_message(BOT_PROFE_TOKEN, chat_id, f"Tus cursos:\n\n{lista}")

    elif text == "/nuevocurso":
        if not profe or not profe.activo:
            await send_message(BOT_PROFE_TOKEN, chat_id, "❌ Necesitas suscripcion activa.")
            return {"ok": True}
        set_estado(db, telegram_id, "paso", "esperando_nombre_curso")
        await send_message(BOT_PROFE_TOKEN, chat_id,
            "✏️ Escribe el nombre y grado del curso:\nEjemplo: <b>Matematicas 9B</b>")

    elif text == "/subirquiz":
        if not profe or not profe.activo:
            await send_message(BOT_PROFE_TOKEN, chat_id, "❌ Necesitas suscripcion activa.")
            return {"ok": True}
        cursos = db.query(Curso).filter(Curso.profe_id == profe.id).all()
        if not cursos:
            await send_message(BOT_PROFE_TOKEN, chat_id, "Primero crea un curso con /nuevocurso")
        else:
            botones = {"inline_keyboard": [[{"text": f"📚 {c.nombre} - {c.grado}", "callback_data": f"curso_{c.id}"}] for c in cursos]}
            await send_message(BOT_PROFE_TOKEN, chat_id, "¿A qué curso pertenece este quiz?", reply_markup=botones)

    elif text == "/enviar":
        if not profe or not profe.activo:
            await send_message(BOT_PROFE_TOKEN, chat_id, "❌ Necesitas suscripcion activa.")
            return {"ok": True}
        cursos = db.query(Curso).filter(Curso.profe_id == profe.id).all()
        if not cursos:
            await send_message(BOT_PROFE_TOKEN, chat_id, "No tienes cursos. Usa /nuevocurso para crear uno.")
        else:
            botones = {"inline_keyboard": [[{
                "text": f"📚 {c.nombre} - {c.grado}",
                "callback_data": f"enviar_curso_{c.id}"
            }] for c in cursos]}
            await send_message(BOT_PROFE_TOKEN, chat_id,
                "📤 ¿A qué curso quieres enviar el archivo?\n\nDespués de seleccionar, envía el archivo.",
                reply_markup=botones)

    elif text and text.startswith("/enviar_a"):
        partes = text.split(" ", 1)
        if len(partes) >= 2:
            try:
                estudiante_id = int(partes[1].strip())
                set_estado(db, telegram_id, "paso", f"enviar_archivo_estudiante_{estudiante_id}")
                await send_message(BOT_PROFE_TOKEN, chat_id,
                    f"📤 Listo. Ahora envía el archivo para el estudiante <code>{estudiante_id}</code>.")
            except:
                await send_message(BOT_PROFE_TOKEN, chat_id,
                    "❌ Formato incorrecto.\nUsa: <code>/enviar_a ID_ESTUDIANTE</code>")
        else:
            await send_message(BOT_PROFE_TOKEN, chat_id,
                "❌ Formato incorrecto.\nUsa: <code>/enviar_a ID_ESTUDIANTE</code>")

    elif text == "/estadisticas":
        if not profe or not profe.activo:
            await send_message(BOT_PROFE_TOKEN, chat_id, "❌ Necesitas suscripcion activa.")
            return {"ok": True}
        resultados = db.query(Resultado).filter(
            Resultado.confirmado == True,
            Resultado.curso_nombre != None,
            Resultado.profe_telegram_id == telegram_id
        ).all()
        if not resultados:
            await send_message(BOT_PROFE_TOKEN, chat_id, "❌ No hay resultados guardados aun.")
        else:
            await send_message(BOT_PROFE_TOKEN, chat_id, "⏳ Generando grafico...")
            grafico = generar_grafico_profe(resultados)
            if grafico:
                await send_photo_bytes(BOT_PROFE_TOKEN, chat_id, grafico,
                    "📊 Estadisticas del grupo — Aprobados vs Reprobados")
            else:
                await send_message(BOT_PROFE_TOKEN, chat_id, "❌ Error generando el grafico.")

    elif text == "/excel" or (text and text.lower().startswith("excel")):
        if not profe or not profe.activo:
            await send_message(BOT_PROFE_TOKEN, chat_id, "❌ Necesitas suscripcion activa.")
            return {"ok": True}
        cursos_con_datos = db.query(Resultado.curso_nombre).filter(
            Resultado.confirmado == True,
            Resultado.curso_nombre != None,
            Resultado.profe_telegram_id == telegram_id
        ).distinct().all()
        if not cursos_con_datos:
            await send_message(BOT_PROFE_TOKEN, chat_id, "❌ No hay resultados guardados aun.")
        else:
            set_estado(db, telegram_id, "paso", "esperando_materia_excel")
            lista = "\n".join([f"• <b>{c[0]}</b>" for c in cursos_con_datos])
            await send_message(BOT_PROFE_TOKEN, chat_id,
                f"📊 ¿De qué materia quieres el Excel?\n\nMaterias disponibles:\n{lista}\n\nEscribe el nombre de la materia:")

    elif text and text.startswith("/responder_voz"):
        partes = text.split(" ", 1)
        if len(partes) >= 2:
            try:
                estudiante_id = int(partes[1].strip())
                set_estado(db, telegram_id, "paso", f"responder_voz_{estudiante_id}")
                await send_message(BOT_PROFE_TOKEN, chat_id,
                    f"🎙️ Listo. Ahora graba y envía tu nota de voz para el estudiante <code>{estudiante_id}</code>.")
            except:
                await send_message(BOT_PROFE_TOKEN, chat_id,
                    "❌ Formato incorrecto.\nUsa: <code>/responder_voz ID_ESTUDIANTE</code>")
        else:
            await send_message(BOT_PROFE_TOKEN, chat_id,
                "❌ Formato incorrecto.\nUsa: <code>/responder_voz ID_ESTUDIANTE</code>")

    elif text and text.startswith("/responder"):
        partes = text.split(" ", 2)
        if len(partes) >= 3:
            try:
                estudiante_chat_id = int(partes[1])
                respuesta = partes[2]
                await send_message(BOT_ESTUDIANTE_TOKEN, estudiante_chat_id,
                    f"📬 <b>Respuesta de tu profe:</b>\n\n{respuesta}")
                await send_message(BOT_PROFE_TOKEN, chat_id, "✅ Respuesta enviada al estudiante.")
            except Exception as e:
                await send_message(BOT_PROFE_TOKEN, chat_id,
                    f"❌ Error: {str(e)}\nFormato: <code>/responder ID_ESTUDIANTE tu respuesta</code>")
        else:
            await send_message(BOT_PROFE_TOKEN, chat_id,
                "❌ Formato incorrecto.\nUsa: <code>/responder ID_ESTUDIANTE tu respuesta aqui</code>")

    elif text and not text.startswith("/"):
        paso = get_estado(db, telegram_id, "paso")

        if paso == "esperando_nombre_curso" and profe and profe.activo:
            partes = text.rsplit(" ", 1)
            nom = partes[0]
            grado = partes[1] if len(partes) > 1 else ""
            nuevo_curso = Curso(id=uuid.uuid4(), profe_id=profe.id, nombre=nom, grado=grado)
            db.add(nuevo_curso)
            del_estado(db, telegram_id, "paso")
            db.commit()
            await send_message(BOT_PROFE_TOKEN, chat_id,
                f"✅ Curso <b>{nom} {grado}</b> creado!\n\nUsa /subirquiz para subir un quiz.")

        elif paso == "esperando_nombre_quiz":
            set_estado(db, telegram_id, "quiz_nombre", text.strip())
            set_estado(db, telegram_id, "paso", "esperando_pdf_zipgrade")
            await send_message(BOT_PROFE_TOKEN, chat_id,
                f"✅ Quiz: <b>{text.strip()}</b>\n\n📎 Ahora envíame el PDF de ZipGrade.")

        elif paso == "esperando_materia_excel":
            materia = text.strip()
            del_estado(db, telegram_id, "paso")
            quizzes = db.query(Resultado.quiz_nombre).filter(
                Resultado.curso_nombre.ilike(f"%{materia}%"),
                Resultado.confirmado == True,
                Resultado.quiz_nombre != None,
                Resultado.profe_telegram_id == telegram_id
            ).distinct().all()
            if not quizzes:
                await send_message(BOT_PROFE_TOKEN, chat_id,
                    f"❌ No encontré resultados para <b>{materia}</b>.")
            else:
                botones_lista = [[{"text": f"📝 {q[0]}", "callback_data": f"excel_quiz_{materia}|{q[0]}"}] for q in quizzes]
                botones_lista.append([{"text": "📊 Todos los quizzes", "callback_data": f"excel_todos_{materia}"}])
                botones = {"inline_keyboard": botones_lista}
                await send_message(BOT_PROFE_TOKEN, chat_id,
                    f"📚 <b>{materia}</b> — ¿De qué quiz quieres el Excel?",
                    reply_markup=botones)

        elif "PAG" in text[:5]:
            resultados_db = db.query(Resultado).filter(
                Resultado.nombre_temp.like("PAG%"),
                Resultado.confirmado == False,
                Resultado.profe_telegram_id == telegram_id
            ).all()
            if not resultados_db:
                await send_message(BOT_PROFE_TOKEN, chat_id,
                    "❌ No encontré el PDF procesado. Por favor vuelve a enviar el PDF primero.")
                return {"ok": True}
            lineas = [l.strip() for l in text.split('\n') if l.strip() and l.strip()[:3] == "PAG"]
            nombres_asignados = 0
            for linea in lineas:
                try:
                    partes = linea.split(":")
                    num_pag = int(partes[0].replace("PAG", "").strip())
                    nombre_real = partes[1].split("-")[0].strip()
                    for r in resultados_db:
                        if r.pagina == num_pag:
                            r.nombre_temp = nombre_real
                            r.confirmado = True
                            nombres_asignados += 1
                            break
                except:
                    continue
            db.commit()
            del_estado(db, telegram_id, "paso")
            del_estado(db, telegram_id, "curso_seleccionado")
            del_estado(db, telegram_id, "quiz_nombre")
            curso_n = resultados_db[0].curso_nombre if resultados_db else ""
            quiz_n = resultados_db[0].quiz_nombre if resultados_db else ""
            resumen = "\n".join([f"• <b>{r.nombre_temp}</b>: {r.nota}/5.0" for r in resultados_db])
            await send_message(BOT_PROFE_TOKEN, chat_id,
                f"✅ <b>{nombres_asignados} estudiantes guardados!</b>\n"
                f"📚 Curso: <b>{curso_n}</b>\n📝 Quiz: <b>{quiz_n}</b>\n\n{resumen}\n\n"
                f"💡 Usa /excel o /estadisticas para ver reportes.")

        else:
            await send_message(BOT_PROFE_TOKEN, chat_id,
                "Comandos:\n/start\n/micursos\n/nuevocurso\n/subirquiz\n/excel\n/estadisticas\n/enviar\n/grupos\n/estado")

    return {"ok": True}

@app.post("/webhook/estudiante")
async def webhook_estudiante(request: Request, db: Session = Depends(get_db)):
    data = await request.json()
    callback = data.get("callback_query", {})
    message = data.get("message", {})

    if callback:
        chat_id = callback.get("from", {}).get("id")
        telegram_id = chat_id
        cb_data = callback.get("data", "")

        if cb_data.startswith("grupo_est_"):
            curso_id = cb_data.replace("grupo_est_", "")
            curso = db.query(Curso).filter(Curso.id == curso_id).first()
            if curso:
                est_nombre = callback.get("from", {}).get("first_name", "Un estudiante")
                entrar_grupo(db, telegram_id, curso_id, f"{curso.nombre} {curso.grado}")
                await send_message(BOT_ESTUDIANTE_TOKEN, chat_id,
                    f"🏫 <b>Entraste al grupo: {curso.nombre} {curso.grado}</b>\n\n"
                    f"Todo lo que escribas o envíes llegará a tus compañeros y al profe.\n\n"
                    f"Usa /salir_grupo para volver a tu chat personal.")
                # ✅ NOTIFICAR AL PROFE que el estudiante entró al grupo
                profe_curso = db.query(Profe).filter(Profe.id == curso.profe_id).first()
                if profe_curso:
                    try:
                        await send_message(BOT_PROFE_TOKEN, profe_curso.telegram_id,
                            f"🔔 <b>{est_nombre} entró al grupo {curso.nombre} {curso.grado}</b>\n\n"
                            f"Usa /grupos para entrar y participar con ellos.")
                    except:
                        pass

        elif cb_data.startswith("duda_materia_"):
            partes = cb_data.replace("duda_materia_", "").split("|", 1)
            materia = partes[0]
            profe_tid = int(partes[1]) if len(partes) > 1 else PROFE_CHAT_ID
            set_estado(db, telegram_id, "duda_materia", materia)
            set_estado(db, telegram_id, "duda_profe_id", str(profe_tid))
            set_estado(db, telegram_id, "esperando_duda", "si")
            set_estado(db, telegram_id, "sesion_inicio", datetime.now().isoformat())
            await send_message(BOT_ESTUDIANTE_TOKEN, chat_id,
                f"📚 Materia: <b>{materia}</b>\n\n"
                f"✏️ Puedes enviar mensajes, notas de voz 🎙️, archivos 📎 e imágenes 🖼️ durante 15 minutos.\n\n"
                f"La sesion se cerrara automaticamente.")
        return {"ok": True}

    chat_id = message.get("chat", {}).get("id")
    text = message.get("text", "")
    telegram_id = message.get("from", {}).get("id")
    nombre = message.get("from", {}).get("first_name", "Estudiante")
    voice = message.get("voice", {})
    photo = message.get("photo", [])
    video = message.get("video", {})
    document = message.get("document", {})

    if not chat_id:
        return {"ok": True}

    estudiante = db.query(Estudiante).filter(Estudiante.telegram_id == telegram_id).first()

    # ── MODO GRUPO ACTIVO DEL ESTUDIANTE ─────────────────────────────────────
    grupo_activo = get_grupo_activo(db, telegram_id)
    if grupo_activo and text != "/salir_grupo" and text != "/grupos" and text != "/duda" and text != "/grafico":
        grupo_nombre = get_grupo_nombre(db, telegram_id)
        nombre_est = estudiante.nombre if estudiante else nombre

        if voice:
            await transmitir_grupo(db, grupo_activo, telegram_id, nombre_est, False,
                "voice", file_id=voice.get("file_id"))
            await send_message(BOT_ESTUDIANTE_TOKEN, chat_id,
                f"🎙️ Nota de voz enviada al grupo <b>{grupo_nombre}</b>.")
            return {"ok": True}

        if photo:
            await transmitir_grupo(db, grupo_activo, telegram_id, nombre_est, False,
                "photo", file_id=photo[-1].get("file_id"), file_name="imagen.jpg")
            await send_message(BOT_ESTUDIANTE_TOKEN, chat_id,
                f"🖼️ Imagen enviada al grupo <b>{grupo_nombre}</b>.")
            return {"ok": True}

        if video:
            fname = video.get("file_name", "video.mp4")
            await transmitir_grupo(db, grupo_activo, telegram_id, nombre_est, False,
                "video", file_id=video.get("file_id"), file_name=fname)
            await send_message(BOT_ESTUDIANTE_TOKEN, chat_id,
                f"🎥 Video enviado al grupo <b>{grupo_nombre}</b>.")
            return {"ok": True}

        if document:
            fname = document.get("file_name", "archivo")
            await transmitir_grupo(db, grupo_activo, telegram_id, nombre_est, False,
                "document", file_id=document.get("file_id"), file_name=fname)
            await send_message(BOT_ESTUDIANTE_TOKEN, chat_id,
                f"📎 Archivo enviado al grupo <b>{grupo_nombre}</b>.")
            return {"ok": True}

        if text and not text.startswith("/"):
            await transmitir_grupo(db, grupo_activo, telegram_id, nombre_est, False,
                "text", text=text)
            return {"ok": True}
    # ─────────────────────────────────────────────────────────────────────────

    if voice:
        esperando = get_estado(db, telegram_id, "esperando_duda")
        if esperando == "si" and sesion_activa(db, telegram_id):
            voice_file_id = voice.get("file_id")
            nombre_est = estudiante.nombre if estudiante else nombre
            materia_duda = get_estado(db, telegram_id, "duda_materia") or "Sin materia"
            profe_id_str = get_estado(db, telegram_id, "duda_profe_id")
            profe_dest = int(profe_id_str) if profe_id_str else PROFE_CHAT_ID
            await send_message(BOT_PROFE_TOKEN, profe_dest,
                f"🎙️ <b>Nota de voz de estudiante:</b>\n"
                f"👤 <b>{nombre_est}</b> | 📚 <b>{materia_duda}</b>\n"
                f"Para responder con texto: <code>/responder {telegram_id} tu respuesta</code>\n"
                f"Para responder con voz: <code>/responder_voz {telegram_id}</code>\n"
                f"Para enviar archivo: <code>/enviar_a {telegram_id}</code>")
            await send_voice(BOT_PROFE_TOKEN, profe_dest, voice_file_id, BOT_ESTUDIANTE_TOKEN)
            await send_message(BOT_ESTUDIANTE_TOKEN, chat_id, "✅ Nota de voz enviada. Puedes seguir enviando mensajes.")
        elif esperando == "si" and not sesion_activa(db, telegram_id):
            del_estado(db, telegram_id, "esperando_duda")
            del_estado(db, telegram_id, "duda_materia")
            del_estado(db, telegram_id, "duda_profe_id")
            del_estado(db, telegram_id, "sesion_inicio")
            await send_message(BOT_ESTUDIANTE_TOKEN, chat_id,
                "⏰ Tu sesion de 15 minutos ha expirado.\n\nUsa /duda para iniciar una nueva sesion.")
        else:
            await send_message(BOT_ESTUDIANTE_TOKEN, chat_id,
                "❌ Primero usa /duda para iniciar una consulta con tu profe.")
        return {"ok": True}

    archivo_recibido = None
    archivo_tipo = None
    archivo_nombre = "archivo"
    if photo:
        archivo_recibido = photo[-1].get("file_id")
        archivo_tipo = "photo"
        archivo_nombre = "imagen.jpg"
    elif video:
        archivo_recibido = video.get("file_id")
        archivo_tipo = "video"
        archivo_nombre = video.get("file_name", "video.mp4")
    elif document and not text:
        archivo_recibido = document.get("file_id")
        archivo_tipo = "document"
        archivo_nombre = document.get("file_name", "archivo")

    if archivo_recibido:
        esperando = get_estado(db, telegram_id, "esperando_duda")
        if esperando == "si" and sesion_activa(db, telegram_id):
            nombre_est = estudiante.nombre if estudiante else nombre
            materia_duda = get_estado(db, telegram_id, "duda_materia") or "Sin materia"
            profe_id_str = get_estado(db, telegram_id, "duda_profe_id")
            profe_dest = int(profe_id_str) if profe_id_str else PROFE_CHAT_ID
            await send_message(BOT_PROFE_TOKEN, profe_dest,
                f"📎 <b>Archivo de estudiante:</b>\n"
                f"👤 <b>{nombre_est}</b> | 📚 <b>{materia_duda}</b>\n"
                f"Para responder: <code>/responder {telegram_id} tu respuesta</code>\n"
                f"Para enviar archivo: <code>/enviar_a {telegram_id}</code>")
            await reenviar_archivo(BOT_PROFE_TOKEN, profe_dest, archivo_recibido,
                BOT_ESTUDIANTE_TOKEN, archivo_nombre, "", archivo_tipo)
            await send_message(BOT_ESTUDIANTE_TOKEN, chat_id, "✅ Archivo enviado a tu profe.")
        elif esperando == "si" and not sesion_activa(db, telegram_id):
            del_estado(db, telegram_id, "esperando_duda")
            del_estado(db, telegram_id, "duda_materia")
            del_estado(db, telegram_id, "duda_profe_id")
            del_estado(db, telegram_id, "sesion_inicio")
            await send_message(BOT_ESTUDIANTE_TOKEN, chat_id,
                "⏰ Tu sesion de 15 minutos ha expirado.\n\nUsa /duda para iniciar una nueva sesion.")
        else:
            await send_message(BOT_ESTUDIANTE_TOKEN, chat_id,
                "❌ Para enviar archivos a tu profe primero usa /duda para iniciar una consulta.")
        return {"ok": True}

    if text == "/start":
        if not estudiante:
            nuevo = Estudiante(id=uuid.uuid4(), telegram_id=telegram_id, nombre=nombre, apellido="", activo=True)
            db.add(nuevo)
            db.commit()
            await send_message(BOT_ESTUDIANTE_TOKEN, chat_id,
                f"👋 Hola <b>{nombre}</b>!\n\nBienvenido al sistema ZipGrade.\n\n"
                f"Puedes:\n• Escribir tu <b>nombre</b> para ver todas tus notas\n"
                f"• Escribir una <b>materia</b> para ver notas de esa materia\n"
                f"• Usar /grafico para ver tu grafico de rendimiento\n"
                f"• Usar /grupos para entrar al chat grupal de una asignatura\n"
                f"• Usar /duda para contactar a tu profe en privado")
        else:
            await send_message(BOT_ESTUDIANTE_TOKEN, chat_id,
                f"✅ Hola <b>{estudiante.nombre}</b>!\n\n"
                f"Comandos:\n/grafico - Ver tu grafico de rendimiento\n"
                f"/grupos - Entrar al chat grupal\n"
                f"/duda - Contactar al profe en privado")

    elif text == "/grupos":
        est = db.query(Estudiante).filter(Estudiante.telegram_id == telegram_id).first()
        if not est or not est.nombre:
            await send_message(BOT_ESTUDIANTE_TOKEN, chat_id,
                "❌ Primero escribe tu nombre completo para registrarte.")
            return {"ok": True}

        cursos_data = db.query(Resultado.curso_nombre, Resultado.profe_telegram_id).filter(
            Resultado.nombre_temp.ilike(f"%{est.nombre}%"),
            Resultado.confirmado == True,
            Resultado.curso_nombre != None
        ).distinct().all()

        if not cursos_data:
            await send_message(BOT_ESTUDIANTE_TOKEN, chat_id,
                "❌ No tienes materias registradas aun. Escribe tu nombre primero.")
            return {"ok": True}

        botones_grupos = []
        for cd in cursos_data:
            curso_nombre = cd[0]
            resultado_sample = db.query(Resultado).filter(
                Resultado.curso_nombre == curso_nombre,
                Resultado.nombre_temp.ilike(f"%{est.nombre}%"),
                Resultado.confirmado == True
            ).first()
            if resultado_sample and resultado_sample.profe_telegram_id:
                profe_obj = db.query(Profe).filter(
                    Profe.telegram_id == resultado_sample.profe_telegram_id
                ).first()
                if profe_obj:
                    curso_obj = db.query(Curso).filter(
                        Curso.profe_id == profe_obj.id,
                        Curso.nombre == curso_nombre
                    ).first()
                    if curso_obj:
                        botones_grupos.append([{
                            "text": f"🏫 {curso_nombre} - {curso_obj.grado}",
                            "callback_data": f"grupo_est_{curso_obj.id}"
                        }])

        grupo_actual = get_grupo_activo(db, telegram_id)
        grupo_nombre_actual = get_grupo_nombre(db, telegram_id)

        if not botones_grupos:
            await send_message(BOT_ESTUDIANTE_TOKEN, chat_id,
                "❌ No encontré grupos disponibles. Asegúrate de que tus cursos estén configurados.")
            return {"ok": True}

        msg = "🏫 <b>Grupos virtuales</b>\n\nSelecciona el grupo al que quieres entrar:"
        if grupo_actual:
            msg = (f"🏫 Actualmente estás en: <b>{grupo_nombre_actual}</b>\n\n"
                   f"Usa /salir_grupo para salir o cambia de grupo:")
        await send_message(BOT_ESTUDIANTE_TOKEN, chat_id, msg,
            reply_markup={"inline_keyboard": botones_grupos})

    elif text == "/salir_grupo":
        grupo_nombre_actual = get_grupo_nombre(db, telegram_id)
        if grupo_nombre_actual:
            salir_grupo(db, telegram_id)
            await send_message(BOT_ESTUDIANTE_TOKEN, chat_id,
                f"✅ Saliste del grupo <b>{grupo_nombre_actual}</b>.\n\nVolviste a tu chat personal.")
        else:
            await send_message(BOT_ESTUDIANTE_TOKEN, chat_id,
                "No estás en ningún grupo actualmente.")

    elif text == "/duda":
        est = db.query(Estudiante).filter(Estudiante.telegram_id == telegram_id).first()
        if not est or not est.nombre:
            await send_message(BOT_ESTUDIANTE_TOKEN, chat_id,
                "❌ Primero escribe tu nombre completo para registrarte.")
            return {"ok": True}
        materias = db.query(
            Resultado.curso_nombre,
            Resultado.profe_telegram_id
        ).filter(
            Resultado.nombre_temp.ilike(f"%{est.nombre}%"),
            Resultado.confirmado == True,
            Resultado.curso_nombre != None,
            Resultado.profe_telegram_id != None
        ).distinct().all()
        if not materias:
            await send_message(BOT_ESTUDIANTE_TOKEN, chat_id,
                "❌ No tienes materias registradas aun. Escribe tu nombre primero.")
        else:
            botones = {"inline_keyboard": [
                [{"text": f"📚 {m[0]}", "callback_data": f"duda_materia_{m[0]}|{m[1]}"}]
                for m in materias
            ]}
            await send_message(BOT_ESTUDIANTE_TOKEN, chat_id,
                "📚 ¿Sobre qué materia tienes la duda?",
                reply_markup=botones)

    elif text == "/grafico":
        est = db.query(Estudiante).filter(Estudiante.telegram_id == telegram_id).first()
        if not est or not est.nombre:
            await send_message(BOT_ESTUDIANTE_TOKEN, chat_id,
                "❌ Primero escribe tu nombre completo para registrarte.")
            return {"ok": True}
        resultados = db.query(Resultado).filter(
            Resultado.nombre_temp.ilike(f"%{est.nombre}%"),
            Resultado.confirmado == True,
            Resultado.curso_nombre != None
        ).all()
        if not resultados:
            await send_message(BOT_ESTUDIANTE_TOKEN, chat_id, "❌ No tienes resultados registrados aun.")
        else:
            await send_message(BOT_ESTUDIANTE_TOKEN, chat_id, "⏳ Generando tu grafico de rendimiento...")
            grafico = generar_grafico_estudiante(est.nombre, resultados)
            if grafico:
                await send_photo_bytes(BOT_ESTUDIANTE_TOKEN, chat_id, grafico,
                    f"📊 Tu rendimiento academico, {est.nombre}")
            else:
                await send_message(BOT_ESTUDIANTE_TOKEN, chat_id, "❌ Error generando el grafico.")

    elif text and not text.startswith("/"):
        esperando = get_estado(db, telegram_id, "esperando_duda")

        if esperando == "si" and sesion_activa(db, telegram_id):
            nombre_est = estudiante.nombre if estudiante else nombre
            materia_duda = get_estado(db, telegram_id, "duda_materia") or "Sin materia"
            profe_id_str = get_estado(db, telegram_id, "duda_profe_id")
            profe_dest = int(profe_id_str) if profe_id_str else PROFE_CHAT_ID
            mins = tiempo_restante(db, telegram_id)
            await send_message(BOT_PROFE_TOKEN, profe_dest,
                f"📩 <b>Mensaje de estudiante:</b>\n"
                f"👤 <b>{nombre_est}</b> | 📚 <b>{materia_duda}</b>\n\n"
                f"💬 {text}\n\n"
                f"Para responder con texto: <code>/responder {telegram_id} tu respuesta</code>\n"
                f"Para responder con voz: <code>/responder_voz {telegram_id}</code>\n"
                f"Para enviar archivo: <code>/enviar_a {telegram_id}</code>")
            await send_message(BOT_ESTUDIANTE_TOKEN, chat_id,
                f"✅ Mensaje enviado. Sesion activa por {mins} min mas.")
        elif esperando == "si" and not sesion_activa(db, telegram_id):
            del_estado(db, telegram_id, "esperando_duda")
            del_estado(db, telegram_id, "duda_materia")
            del_estado(db, telegram_id, "duda_profe_id")
            del_estado(db, telegram_id, "sesion_inicio")
            await send_message(BOT_ESTUDIANTE_TOKEN, chat_id,
                "⏰ Tu sesion de 15 minutos ha expirado.\n\nUsa /duda para iniciar una nueva sesion.")
        else:
            busqueda = text.strip()
            resultados_materia = db.query(Resultado).filter(
                Resultado.curso_nombre.ilike(f"%{busqueda}%"),
                Resultado.confirmado == True
            ).all()
            resultados_nombre = db.query(Resultado).filter(
                Resultado.nombre_temp.ilike(f"%{busqueda}%"),
                Resultado.confirmado == True
            ).all()

            if resultados_materia and not resultados_nombre:
                est = db.query(Estudiante).filter(Estudiante.telegram_id == telegram_id).first()
                if est and est.nombre:
                    resultados = db.query(Resultado).filter(
                        Resultado.curso_nombre.ilike(f"%{busqueda}%"),
                        Resultado.nombre_temp.ilike(f"%{est.nombre}%"),
                        Resultado.confirmado == True
                    ).all()
                    if not resultados:
                        await send_message(BOT_ESTUDIANTE_TOKEN, chat_id,
                            f"❌ No encontré tus notas en <b>{busqueda}</b>.")
                    else:
                        msg = f"📚 <b>Tus notas en {busqueda.title()}:</b>\n\n"
                        for r in resultados:
                            msg += f"📝 <b>{r.quiz_nombre}</b>: {r.nota}/5.0 ({r.porcentaje}%)\n"
                        await send_message(BOT_ESTUDIANTE_TOKEN, chat_id, msg)
                        for r in resultados:
                            if r.imagen_url:
                                await send_photo(BOT_ESTUDIANTE_TOKEN, chat_id, r.imagen_url,
                                    f"📋 {r.quiz_nombre} - Tu hoja de respuestas")
                            if r.quiz_pdf_url:
                                await send_document_url(BOT_ESTUDIANTE_TOKEN, chat_id, r.quiz_pdf_url,
                                    f"📄 {r.quiz_nombre} - PDF del quiz")
                else:
                    await send_message(BOT_ESTUDIANTE_TOKEN, chat_id,
                        f"❌ Primero escribe tu nombre completo para registrarte.")

            elif resultados_nombre:
                est = db.query(Estudiante).filter(Estudiante.telegram_id == telegram_id).first()
                if est and est.nombre != busqueda:
                    est.nombre = busqueda
                    db.commit()
                msg = f"📊 <b>Todas tus notas ({busqueda}):</b>\n\n"
                for r in resultados_nombre:
                    curso = r.curso_nombre or "Sin curso"
                    quiz = r.quiz_nombre or "Sin quiz"
                    msg += f"📚 <b>{curso}</b> - {quiz}: <b>{r.nota}/5.0</b> ({r.porcentaje}%)\n"
                await send_message(BOT_ESTUDIANTE_TOKEN, chat_id, msg)
                for r in resultados_nombre:
                    if r.imagen_url:
                        await send_photo(BOT_ESTUDIANTE_TOKEN, chat_id, r.imagen_url,
                            f"📋 {r.curso_nombre} - {r.quiz_nombre}")
                    if r.quiz_pdf_url:
                        await send_document_url(BOT_ESTUDIANTE_TOKEN, chat_id, r.quiz_pdf_url,
                            f"📄 {r.curso_nombre} - {r.quiz_nombre}")
            else:
                await send_message(BOT_ESTUDIANTE_TOKEN, chat_id,
                    f"❌ No encontré resultados para <b>{busqueda}</b>.\n\n"
                    f"Intenta con tu nombre completo o el nombre de la materia.\n"
                    f"¿Tienes una duda? Usa /duda para contactar a tu profe.")

    return {"ok": True}

@app.post("/profes/registrar")
def registrar_profe(data: dict, db: Session = Depends(get_db)):
    profe = db.query(Profe).filter(Profe.telegram_id == data["telegram_id"]).first()
    if profe:
        return {"id": str(profe.id), "nombre": profe.nombre, "activo": profe.activo}
    nuevo = Profe(id=uuid.uuid4(), telegram_id=data["telegram_id"], nombre=data.get("nombre", ""), email="", activo=False)
    db.add(nuevo)
    db.commit()
    return {"id": str(nuevo.id), "nombre": nuevo.nombre, "activo": nuevo.activo}

@app.get("/profes/by-telegram/{telegram_id}")
def get_profe_by_telegram(telegram_id: int, db: Session = Depends(get_db)):
    profe = db.query(Profe).filter(Profe.telegram_id == telegram_id).first()
    if not profe:
        raise HTTPException(status_code=404, detail="Profe no encontrado")
    return {"id": str(profe.id), "nombre": profe.nombre, "activo": profe.activo}

@app.get("/profes/activo/{telegram_id}")
def check_profe_activo(telegram_id: int, db: Session = Depends(get_db)):
    return {"activo": profe_activo(telegram_id, db)}

@app.post("/estudiantes/registrar")
def registrar_estudiante(data: dict, db: Session = Depends(get_db)):
    est = db.query(Estudiante).filter(Estudiante.telegram_id == data["telegram_id"]).first()
    if est:
        return {"id": str(est.id), "nombre": est.nombre, "activo": est.activo}
    nuevo = Estudiante(id=uuid.uuid4(), telegram_id=data["telegram_id"], nombre=data.get("nombre", ""), apellido="", activo=False)
    db.add(nuevo)
    db.commit()
    return {"id": str(nuevo.id), "nombre": nuevo.nombre, "activo": nuevo.activo}

@app.get("/estudiantes/by-telegram/{telegram_id}")
def get_estudiante_by_telegram(telegram_id: int, db: Session = Depends(get_db)):
    est = db.query(Estudiante).filter(Estudiante.telegram_id == telegram_id).first()
    if not est:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    return {"id": str(est.id), "nombre": est.nombre, "apellido": est.apellido, "activo": est.activo}

@app.get("/estudiantes/activo/{telegram_id}")
def check_estudiante_activo(telegram_id: int, db: Session = Depends(get_db)):
    return {"activo": estudiante_activo(telegram_id, db)}

@app.post("/cursos/crear")
def crear_curso(data: dict, db: Session = Depends(get_db)):
    nuevo = Curso(id=uuid.uuid4(), profe_id=data["profe_id"], nombre=data["nombre"], grado=data.get("grado", ""))
    db.add(nuevo)
    db.commit()
    return {"id": str(nuevo.id), "nombre": nuevo.nombre, "grado": nuevo.grado}

@app.get("/cursos/by-profe-telegram/{telegram_id}")
def cursos_by_profe(telegram_id: int, db: Session = Depends(get_db)):
    profe = db.query(Profe).filter(Profe.telegram_id == telegram_id).first()
    if not profe:
        return []
    cursos = db.query(Curso).filter(Curso.profe_id == profe.id).all()
    return [{"id": str(c.id), "nombre": c.nombre, "grado": c.grado} for c in cursos]

@app.post("/quizzes/procesar-pdf")
async def procesar_pdf_endpoint(archivo: UploadFile = File(...), curso_id: str = Form(...), db: Session = Depends(get_db)):
    contenido = await archivo.read()
    resultados = await procesar_pdf_zipgrade(contenido)
    return {"resultados": resultados, "total": len(resultados)}

@app.get("/resultados/historial/{estudiante_id}")
def historial_estudiante(estudiante_id: str, db: Session = Depends(get_db)):
    return db.query(Resultado).filter(
        Resultado.estudiante_id == estudiante_id,
        Resultado.confirmado == True).all()

@app.post("/admin/activar-profe/{telegram_id}")
def admin_activar_profe(telegram_id: int, db: Session = Depends(get_db)):
    activar_profe(telegram_id, db)
    return {"ok": True}

@app.post("/admin/desactivar-profe/{telegram_id}")
def admin_desactivar_profe(telegram_id: int, db: Session = Depends(get_db)):
    desactivar_profe(telegram_id, db)
    return {"ok": True}

@app.post("/admin/activar-estudiante/{telegram_id}")
def admin_activar_estudiante(telegram_id: int, db: Session = Depends(get_db)):
    activar_estudiante(telegram_id, db)
    return {"ok": True}

@app.post("/admin/desactivar-estudiante/{telegram_id}")
def admin_desactivar_estudiante(telegram_id: int, db: Session = Depends(get_db)):
    desactivar_estudiante(telegram_id, db)
    return {"ok": True}